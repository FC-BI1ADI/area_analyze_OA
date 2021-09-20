import sys, os, shutil, time, datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd

sys.path.append(r"D:\PROC\Projects\STD_CODE")

from excel_fc import *
from dir_file_fc import *
from openpyxl_fc import *
import code_segment as CS


PASSWORD = "xsgl2020"

# 程序运行环境设置
# 设置DataFrame显示行宽
pd.set_option('Display.width', 5000)
# 设置DataFrame最大显示列数
pd.set_option('Display.max_columns', 60)

# 获取TODO目录下的文件名列表todo_files作为待处理文件列表
todo_files = get_filenames(r".\TODO")
# 对TODO列表进行清理，删除非XLSX文件和分析过程文件
# ---------------------------------------------
temp_files = []
for i in range(0, len(todo_files)):
    if todo_files[i].find(".xlsx") != -1 and todo_files[i] != "区域项目报备记录汇总表.xlsx":
        # 文件中未含.xlsx后缀
        temp_files.append(os.path.join(os.getcwd(), "TODO", todo_files[i]))
todo_files = temp_files
# 输出需要处理的文件名列表
print("--------- ToDo Files ---------")
for f in todo_files:
    print("ToDo:[%s]"%f)
print("========= ToDo Files =========")
# 清除待处理文件的密码，并且将文件单元格进行拆分
for f in todo_files:
    # 清除excel文件密码
    print(">>> 正在清除 [%s] 的密码" % (f))
    remove_xlsx_password(f, PASSWORD)
    demerge_excel_cell(f)

# 处理所有待处理文件，汇总成为区域项目报德记录汇总表.xlsx
# ---------------------------------------------
# 保留表头，清空区域项目报备记录汇总表所有记录（从第3行起删除）
# file_gather = os.path.join(os.getcwd(), "DATA", "区域项目报备记录汇总表.xlsx")
wb_gather = openpyxl.load_workbook(r".\DATA\区域项目报备记录汇总表.xlsx")
ws_gather = wb_gather.active

print(">>> 汇总提交的区域项目报备分表...")
# 清空区域项目报备记录汇总表的原有记录，供保留表头
# 从最大行数倒着删除到第3行，防止正向删除导致的跳动
for i in range(ws_gather.max_row, 2, -1):
    ws_gather.delete_rows(i)

# 扫描每个文件，将数据记入汇总表
for f in todo_files:
    wb_f = openpyxl.load_workbook(f, data_only=True)
    # 非Pipeline项目
    ws_np = wb_f["非Pipeline项目"]
    for i in range(3, ws_np.max_row + 1):
        row = get_1row(ws_np, i)
        row.append("非Pipeline项目")
        # 若扫描到项目编号为空，则停止扫描
        if row[1] == None:
            break
        else:
            ws_gather.append(row)
    # Pipeline项目
    ws_p = wb_f["Pipeline项目"]
    for i in range(3, ws_p.max_row + 1):
        row = get_1row(ws_p, i)
        row.append("Pipeline项目")
        # 若扫描到项目编号为空，则停止扫描
        if row[1] == None:
            break
        else:
            ws_gather.append(row)
    # 关闭扫描文件
    wb_f.close()
# 保存对区域项目报备记录汇总表.xlsx进行的修改
wb_gather.save(r".\DATA\区域项目报备记录汇总表.xlsx")
# 至此，将TODO目录下所有XLSX文件汇总到[区域项目报备记录汇总表.xlsx]中
print(">>> 区域项目报备分表汇总完成")
# ---------------------------------------------

# 将汇总表和分析表分别读入dataframe中，df_gather / df_analyze，为DF保留表头
df_gather = pd.read_excel(r".\DATA\区域项目报备记录汇总表.xlsx", header=1)
df_analyze = pd.read_excel(r".\DATA\区域项目报备分析表.xlsx", header=1)

# 汇总表和分析表读入后，进行数据清洗
# 考虑加入gather表的去重

# 获取当前时间作为分析时间
analyze_time = time.localtime(time.time())

print(">>> 比对汇总表和以往的分析表...")
# 以汇总表和分析表的索引进行嵌套循环，扫描analyze表中是否有gather表的记录，若有则确定是否修改
for idx_gather in range(df_gather.shape[0]):
    # 设置查找标志为假，默认为未找到
    found_flag = False
    for idx_analyze in range(df_analyze.shape[0]):
        # 若项目编号相同，则比较三联(tree uion,简写tu_) 7、8、9
        if df_gather.iloc[idx_gather, 1] == df_analyze.iloc[idx_analyze, 1]:
            # 项目编号相同，获取gather表和analyze表的三联TU
            tu_gather = "%s/%s/%s" % (
                df_gather.iloc[idx_gather, 7], df_gather.iloc[idx_gather, 8], df_gather.iloc[idx_gather, 9])
            tu_analyze = "%s/%s/%s" % (
                df_analyze.iloc[idx_analyze, 7], df_analyze.iloc[idx_analyze, 8], df_analyze.iloc[idx_analyze, 9])
            # 比较三联，若相同则比较存货数量，若相同则状态为不变，否则为修改
            if tu_gather == tu_analyze:
                # 若项目编号相同，且三联相同，则判断为找到相同记录，修改标志位为真
                found_flag = True
                if df_gather.iloc[idx_gather, 11] == df_analyze.iloc[idx_analyze, 11]:
                    df_analyze.iloc[idx_analyze, 19] = "未变"
                    df_analyze.iloc[idx_analyze, 20] = ""
                else:
                    # df中第11列为存货数量
                    if pd.isna(df_analyze.iloc[idx_analyze, 11]) or pd.isna(df_gather.iloc[idx_gather, 11]):
                        # analyze表为空，gather表不为空
                        if pd.isna(df_analyze.iloc[idx_analyze, 11]) == True and pd.isna(df_gather.iloc[idx_gather, 11]) == False:
                            df_analyze.iloc[idx_analyze, 19] = "修改"
                            df_analyze.iloc[idx_analyze, 20] = "数量:0-->%d"%(df_gather.iloc[idx_gather, 11])
                            # 更新存货数量
                            df_analyze.iloc[idx_analyze, 11] = df_gather.iloc[idx_gather, 11]
                        # analyze表不为空，gather表为空
                        if pd.isna(df_analyze.iloc[idx_analyze, 11]) == False and pd.isna(df_gather.iloc[idx_gather, 11]) == True:
                            df_analyze.iloc[idx_analyze, 19] = "修改"
                            df_analyze.iloc[idx_analyze, 20] = "数量:%d-->0"%(df_analyze.iloc[idx_analyze, 11])
                        # analyze表、gather表均为空
                        if pd.isna(df_analyze.iloc[idx_analyze, 11]) == True and pd.isna(df_gather.iloc[idx_gather, 11]) == True:
                            df_analyze.iloc[idx_analyze, 19] = "未变"
                            df_analyze.iloc[idx_analyze, 20] = "数量为空"
                    else:
                        df_analyze.iloc[idx_analyze, 19] = "修改"
                        df_analyze.iloc[idx_analyze, 20] = "数量:%d-->%d" % (df_analyze.iloc[idx_analyze, 11], df_gather.iloc[idx_gather, 11])
                        # 更新存货数量
                        df_analyze.iloc[idx_analyze, 11] = df_gather.iloc[idx_gather, 11]
    # 扫描分析表完毕，根据查找标志位判断是否将汇总表中记录添加到分析表中
    if found_flag == False:
        # 复制gather表中当前记录，并添加到analyze表中
        # 初始化row列表，然后构建一行数据
        row = []
        for i in range(19):  # 复制索引为0-18的共19列数据至row列表中
            row.append(df_gather.iloc[idx_gather, i])
        # 设置第19列变动类型：新增、修改、删除
        row.append("新增")
        # 设置第20列对比分析
        row.append("")
        # 设置第21列项目类型：Pipeline项目/非Pipeline项目
        row.append(df_gather.iloc[idx_gather, 19])
        # 设置第22列项目状态：【暂时PASS】
        row.append("")
        # 设置第23列项目阶段变动时间：【暂时PASS】
        row.append(pd.to_datetime(time.strftime('%Y%m%d', time.localtime(time.time()))))
        # 至此，构建完成一行完整24列的数据列表
        row_series = pd.Series(row, index=df_analyze.columns)
        df_analyze = df_analyze.append(row_series, ignore_index=True)

# 以汇总表和分析表的索引进行嵌套循环，扫描gather表中是否有analyze表中存在的记录，若有则确定有记录被删除
for idx_analyze in range(df_analyze.shape[0]):
    # 设置查找标志为假，默认为未找到
    found_flag = False
    for idx_gather in range(df_gather.shape[0]):
        if df_gather.iloc[idx_gather, 1] == df_analyze.iloc[idx_analyze, 1]:
            # 项目编号相同，获取gather表和analyze表的三联TU
            tu_gather = "%s/%s/%s" % (
                df_gather.iloc[idx_gather, 7], df_gather.iloc[idx_gather, 8], df_gather.iloc[idx_gather, 9])
            tu_analyze = "%s/%s/%s" % (
                df_analyze.iloc[idx_analyze, 7], df_analyze.iloc[idx_analyze, 8], df_analyze.iloc[idx_analyze, 9])
            # 比较三联，若相同则修改标志位为真，即找到相同记录
            if tu_gather == tu_analyze:
                found_flag = True
                # 判断项目变动状态，进行预警3个月、小僵尸6个月、大僵尸12个月标注（项目状态）
                # 如果汇总表和分析表的项目阶段没变化，且当前时间与项目阶段变动时间差值大于固定时间，进行项目状态修改
                # --------------------------------------------------------------
                if df_gather.iloc[idx_gather, 13] == df_analyze.iloc[idx_analyze, 13]:
                    interval = (datetime.datetime.now() - df_analyze.iloc[idx_analyze, 23])
                    if interval.days <= 90:
                        df_analyze.iloc[idx_analyze, 22] = "正常"
                    if interval.days > 90:
                        df_analyze.iloc[idx_analyze, 22] = "预警"
                    if interval.days > 182:
                        df_analyze.iloc[idx_analyze, 22] = "小僵尸"
                    if interval.days > 365:
                        df_analyze.iloc[idx_analyze, 22] = "大僵尸"
                else:
                    # 若分析时发现项目阶段发生变化，即修改23列项目阶段变动时间
                    df_analyze.iloc[idx_analyze, 13] = df_gather.iloc[idx_gather, 13]
                    df_analyze.iloc[idx_analyze, 23] = pd.to_datetime(
                        time.strftime('%Y%m%d', time.localtime(time.time())))

    # 如果分析表中记录未在汇总表中找到，即判断为已删除
    if found_flag == False:
        df_analyze.iloc[idx_analyze, 19] = "删除"
        df_analyze.iloc[idx_analyze, 20] = ""
# 对df_analyze按项目编号进行排序
df_analyze.sort_values(by='项目编码', ascending=True)

print(">>> 输出区域项目报备分析表...")
# 输出报备分析DF数据至分析表EXCEL文件中
writer = pd.ExcelWriter(r".\DATA\区域项目报备分析表.xlsx", datetime_format='YYYY-MM-DD')
df_analyze.to_excel(writer, sheet_name='Analyze', index=False)
writer.close()

#######################################################################################################################
# 按照定制要求，设置输出EXCEL格式
# 打开待设置分析表文件，获得操作SHEET
print(">>> 调整区域项目报备分析表格式...")
wb = openpyxl.load_workbook(r".\DATA\区域项目报备分析表.xlsx")
ws = wb.active
# 设置行高
ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 20
# 设置列宽
#               A   B   C   D   E   F   G   H   I   J   K  L   M   N   O   P  Q    R   S   T   U   V   W   X
column_width = [5, 15, 30, 20, 10, 20, 10, 15, 20, 25, 15, 8, 10, 15, 15, 15, 10, 15, 20, 10, 20, 15, 20, 20]
for i in range(1, ws.max_column + 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width[i - 1]

# 增加首行标题行
ws.insert_rows(1)
# 设置首行标题为区域项目报备分析
ws['A1'] = "区域项目报备分析"
ws.merge_cells("A1:Q1")
ws['W1'] = "分析时间"
ws['X1'] = CS.get_current_timestr()

# 设置表头格式（字体、背景色）
title_font = Font(name=u'微软雅黑', bold=True, size=18)
header_font = Font(name=u'宋体', bold=True, size=12)
mark_font = Font(name=u'宋体', bold=True, size=12)
delete_font = Font(name=u'宋体', color='FF0000', strike=True)
modify_font = Font(name=u'宋体', color='4682B4')
newadd_font = Font(name=u'宋体', color='8B4C39')

align_center = Alignment(horizontal='center', wrap_text=True)
align_left = Alignment(horizontal='left', wrap_text=True)
align_right = Alignment(horizontal='right', wrap_text=True)

# 设置表头标题行
ws['A1'].font = title_font
ws['A1'].fill = PatternFill("solid", fgColor='C8C8C8')
ws['A1'].alignment = align_center
# 设置表头分析时间
ws['W1'].font = mark_font
# 设置表头Header行字体和背景色
for i in range(1, 25):
    ws.cell(2, i).font = header_font
    if 1 <= i <= 3:
        ws.cell(2, i).fill = PatternFill("solid", fgColor='87CEFA')
    if 4 <= i <= 5:
        ws.cell(2, i).fill = PatternFill("solid", fgColor='98FB98')
    if 6 <= i <= 7:
        ws.cell(2, i).fill = PatternFill("solid", fgColor='1E90FF')
    if 8 <= i <= 10:
        ws.cell(2, i).fill = PatternFill("solid", fgColor='FF82AB')
    if 11 <= i <= 12:
        ws.cell(2, i).fill = PatternFill("solid", fgColor='FFE4E1')
    if i == 13:
        ws.cell(2, i).fill = PatternFill("solid", fgColor='4682B4')
    if 14 <= i <= 16:
        ws.cell(2, i).fill = PatternFill("solid", fgColor='87CEFA')
    if i == 17:
        ws.cell(2, i).fill = PatternFill("solid", fgColor='696969')
    if 18 <= i <= 19:
        ws.cell(2, i).fill = PatternFill("solid", fgColor='FF7F24')
    if 20 <= i <= 24:
        ws.cell(2, i).fill = PatternFill("solid", fgColor='C1FFC1')
# 扫描分析状态，设置对应格式
for idx_row in range(3, ws.max_row + 1):
    # 设置新增项的提醒色和格式
    if ws.cell(idx_row, 20).value == "新增":
        for j in range(8, 11):
            ws.cell(idx_row, j).font = newadd_font
        ws.cell(idx_row, 20).font = newadd_font
    # 设置删除项的提醒色和格式
    if ws.cell(idx_row, 20).value == "修改":
        for j in range(8, 11):
            ws.cell(idx_row, j).font = modify_font
        ws.cell(idx_row, 20).font = modify_font
    # 设置删除项的提醒色和格式
    if ws.cell(idx_row, 20).value == "删除":
        for j in range(8, 11):
            ws.cell(idx_row, j).font = delete_font
        ws.cell(idx_row, 20).font = delete_font
    # 设置不同状态的提醒色
    if ws.cell(idx_row, 23).value == "预警":
        for j in range(1, 25):
            ws.cell(idx_row, j).fill = PatternFill("solid", fgColor='F9F994')
    if ws.cell(idx_row, 23).value == "小僵尸":
        for j in range(1, 25):
            ws.cell(idx_row, j).fill = PatternFill("solid", fgColor='FFBBFF')
    if ws.cell(idx_row, 23).value == "大僵尸":
        for j in range(1, 25):
            ws.cell(idx_row, j).fill = PatternFill("solid", fgColor='FF4500')
# 保存格式设置修改
wb.save(filename=r".\DATA\区域项目报备分析表.xlsx")

print(">>> 备份区域项目报备分析表至BAK目录...")
# 将当前输出的分析文件加时间戳备份到BAK目录下
analyze_filename = ".\BAK\区域项目报备分析表(%s).xlsx" % (CS.get_current_timeid())
shutil.copyfile(r".\DATA\区域项目报备分析表.xlsx", analyze_filename)
print(">>> 区域项目报备分析全部完成！可在DATA目录下【区域项目报备分析表.xlsx】查看分析结果")