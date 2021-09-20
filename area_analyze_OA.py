import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
import os.path, time


# 使用PANDAS进行数据比对处理，设置对应的状态位

# 将前期表（Previous table)和分析表(Analyze table)分别读入dataframe中，df_p / df_a，为DF保留表头
df_p = pd.read_excel(r".\DATA\previous.xlsx", header=0)
df_a = pd.read_excel(r".\DATA\analyze.xlsx", header=0)

# 前期表和分析表读入后，进行数据清洗


# 输出已经数据清洗后的DATAFRAME
# print("previous.xlsx文件内容如下：")
# print(df_p.head(10))
# print("analyze.xlsx文件内容如下：")
# print(df_a.head(10))


# 获取当前时间作为分析时间
# analyze_time = time.localtime(time.time())
# 获取前期表和分析表的文件修改时间作为判定状态变化的时间标量
time_p = os.path.getmtime(r".\DATA\previous.xlsx")
time_a = os.path.getmtime(r".\DATA\analyze.xlsx")
print("previous.xlsx修改时间为：%s"%(time.ctime(time_p)))
print("analyze.xlsx修改时间为：%s"%(time.ctime(time_a)))
delta_time = time_a - time_p
print(delta_time)
# 需求备注：“预警”：60天；“小僵尸”：180天；“大僵尸”：365天。
if 0 < delta_time < 60*86400:
    warning = "正常"
elif 60*86400 <= delta_time < 180*86400:
    warning = "预警"
elif 180*86400 <= delta_time <365*86400:
    warning = "小僵尸"
elif 365*86400 <= delta_time:
    warning = "大僵尸"
print("Warning FLAG:%s"%(warning))


print(">>> 比对前期表previous.xlsx和分析表analyze.xlsx...")
# 以分析表为基础创建result数据表，简写为df_r
df_r = df_a
# 为结果表df_r创建分析结果列
df_r["分析结果"] = ""


# 以前期表和分析表的索引进行嵌套循环，扫描analyze表中是否有previous表的记录，若有则确定是否修改
# 外部遍历df_p，循环扫描df_a,扫描均以索引方式进行
# 获取项目编号、产品编码的列号
project_no = df_p.columns.get_loc("项目编码")
product_no = df_p.columns.get_loc("产品编码")
product_count = df_p.columns.get_loc("数量")
project_stage = df_p.columns.get_loc("项目推进阶段")
product_amount = df_p.columns.get_loc("预计落单金额")
product_outtime = df_p.columns.get_loc("预计出库时间")
analyze_result = df_r.columns.get_loc("分析结果")

# 以遍历前期表为基础，扫描分析表
for idx_p in range(df_p.shape[0]):
    # 设置查找标志为假，默认为未找到
    found_flag = False

    for idx_a in range(df_a.shape[0]):
        # 若项目编号相同，则比较产品编码
        if df_p.iloc[idx_p, project_no] == df_a.iloc[idx_a, project_no]:
            # 比较previous表和analyze表的产品编码，若相同则比较存货数量，若相同则状态为不变，否则为修改
            if df_p.iloc[idx_p,product_no] == df_a.iloc[idx_a,product_no]:
                # 若项目编号相同，且产品编码相同，则判断为找到相同记录，修改标志位为真
                found_flag = True
                # 依次比较数量、项目推进阶段、预计落单金额、预计出库时间4项内容，有任何一项变化均为有变化
                if df_p.iloc[idx_p, product_count] != df_a.iloc[idx_a, product_count]:
                    df_r.iloc[idx_a, analyze_result] += "#数量：%d-->%d"%(df_p.iloc[idx_p, product_count],df_a.iloc[idx_a, product_count])
                elif df_p.iloc[idx_p,project_stage] != df_a.iloc[idx_a, project_stage]:
                    df_r.iloc[idx_a, analyze_result] += "#阶段：%s-->%s"%(df_p.iloc[idx_p,project_stage],df_a.iloc[idx_a, project_stage])
                elif df_p.iloc[idx_p,product_amount] != df_a.iloc[idx_a, product_amount]:
                    df_r.iloc[idx_a,analyze_result] += "#金额：%d-->%d"%(df_p.iloc[idx_p,product_amount],df_a.iloc[idx_a, product_amount])
                elif df_p.iloc[idx_p,product_outtime] != df_a.iloc[idx_a, product_outtime]:
                    df_r.iloc[idx_a,analyze_result] += "#出库：%s-->%s"%(df_p.iloc[idx_p,product_outtime],df_a.iloc[idx_a, product_outtime])
                else: # 没有变化
                    df_r.iloc[idx_a,analyze_result] = "未变化"

    # 扫描分析表完毕，根据查找标志位判断df_p表中的记录是否在df_a中，若不在则说明记录已删除，需要添加到结果表中并做删除标记
    if found_flag == False:
        # 复制previous表中当前记录，并添加到analyze表中
        #创建交换df
        df_swap = df_p.iloc[[idx_p]]
        df_swap["分析结果"] = "删除"
        df_r = df_r.append(df_swap)

# 以遍历分析表为基础，扫描前期表
for idx_a in range(df_a.shape[0]):
    # 设置查找标志为假，默认为未找到
    found_flag = False
    for idx_p in range(df_p.shape[0]):
        if df_a.iloc[idx_a, project_no] == df_p.iloc[idx_p, project_no]:
            # 比较previous表和analyze表的产品编码，若相同则比较存货数量
            if df_a.iloc[idx_a, product_no] == df_p.iloc[idx_p, product_no]:
            # 若项目编号相同，且产品编码相同，则判断为找到相同记录，修改标志位为真
                found_flag = True
    if found_flag == False:
        df_r.iloc[idx_a,analyze_result] = "新增"

# 对df_r按项目编号进行排序
df_r.sort_values(by="项目编码", ascending=True)

# 从分析结果DataFrame中输出分析结果
print(">>> 输出区域项目分析结果表result.xlsx...")
# 输出报备分析DF数据至分析表EXCEL文件中
writer = pd.ExcelWriter(r".\DATA\result.xlsx", datetime_format='YYYY-MM-DD')
df_r.to_excel(writer, sheet_name='Analyze', index=False)
writer.close()

# 按照定制要求，设置输出EXCEL格式
# 打开待设置结果表文件，获得操作SHEET
print(">>> 调整区域项目报备分析结果表格式...")
wb = openpyxl.load_workbook(r".\DATA\result.xlsx")
ws = wb.active
# 设置行高
ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 20
# 设置列宽
#               A   B   C   D   E   F   G   H   I   J   K   L   M   N   O   P   Q   R   S   T   U   V   W   X
column_width = [15, 8, 15, 15, 25, 20, 20, 10,  8, 15, 15, 15, 12,  8, 10, 10,  8, 10, 10, 15, 15, 15,  8, 30]
for i in range(1, ws.max_column + 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width[i - 1]

# 增加首行标题行
ws.insert_rows(1)
# 设置首行标题为区域项目报备分析
ws['A1'] = "区域项目报备分析"
ws.merge_cells("A1:X1")


# 设置表头格式（字体、背景色）
title_font = Font(name=u'微软雅黑', bold=True, size=18)
header_font = Font(name=u'宋体', bold=True, size=12)
mark_font = Font(name=u'宋体', bold=True, size=12)
delete_font = Font(name=u'宋体', color='FF0000', strike=True) # 红色
modify_font = Font(name=u'宋体', color='336666') #
newadd_font = Font(name=u'宋体', color='0033FF')

align_center = Alignment(horizontal='center', wrap_text=True)
align_left = Alignment(horizontal='left', wrap_text=True)
align_right = Alignment(horizontal='right', wrap_text=True)

# 设置表头标题行
ws['A1'].font = title_font
ws['A1'].fill = PatternFill("solid", fgColor='DDDDDD')
ws['A1'].alignment = align_center

# 设置表头Header行字体和背景色
for i in range(1, 25):
    ws.cell(2, i).font = header_font
    if 1 <= i <= 23:
        ws.cell(2, i).fill = PatternFill("solid", fgColor='0099FF') # 蓝色
    if i == 24:
        ws.cell(2, i).fill = PatternFill("solid", fgColor='00FF66') # 绿色

# 扫描分析状态，设置对应格式
for idx_row in range(3, ws.max_row + 1):
    # 设置第22列，预计出库时间的日期显示格式
    ws.cell(idx_row,22).number_format = "YYYY-MM-DD"
    # 设置新增项的提醒色和格式
    if ws.cell(idx_row, 24).value.find("新增") != -1:
        for j in range(1, ws.max_column + 1):
            ws.cell(idx_row, j).font = newadd_font
    # 设置修改项的提醒色和格式
    if ws.cell(idx_row, 24).value.find("#") != -1:
        for j in range(1, ws.max_column + 1):
            ws.cell(idx_row, j).font = modify_font
    # 设置删除项的提醒色和格式
    if ws.cell(idx_row, 24).value.find("删除") != -1:
        for j in range(1, ws.max_column + 1):
            ws.cell(idx_row, j).font = delete_font

    # 设置不同状态的提醒色
    # 叠加预警标志
    if ws.cell(idx_row, 24).value.find("未变化") != -1:
        ws.cell(idx_row, 24).value += "-%s"%(warning)
    # 根据标志设置格式
    if ws.cell(idx_row, 24).value.find("预警") != -1:
        for j in range(1, ws.max_column +1):
            ws.cell(idx_row, j).fill = PatternFill("solid", fgColor='FF33FF') # 紫色
    if ws.cell(idx_row, 24).value.find("小僵尸") != -1:
        for j in range(1, ws.max_column +1):
            ws.cell(idx_row, j).fill = PatternFill("solid", fgColor='FFCC00') # 橙色
    if ws.cell(idx_row, 24).value.find("大僵尸") != -1:
        for j in range(1, ws.max_column +1):
            ws.cell(idx_row, j).fill = PatternFill("solid", fgColor='CC0000') # 暗红


# 保存格式设置修改
wb.save(filename=r".\DATA\result.xlsx")

print(">>> 区域项目报备分析全部完成,可在DATA目录下【result.xlsx】查看分析结果!")